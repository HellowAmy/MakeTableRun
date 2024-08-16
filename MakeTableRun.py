import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import pandas
import numpy
import sys
import shutil
import os



"""
    解析xlsx文件的每一个sheet的总数量
    并统计出货和未出货的数量
"""
class xlsx_sum_a:

    # 计算结果的数据结构体
    class data_ret:
        def __init__(self):
            self.sheet_name = ""
            self.err_tips = ""
            self.ret_sum = 0.0
            self.ret_sum_empty = 0.0
            self.ret_sum_date = 0.0
            self.ret_count_empty = 0

    # 计算结果的数据结构体
    class data_config:
        def __init__(self):
            self.srcpath = ""
            self.dstpath = ""

    # 创建结果表格
    def __init__(self):
        self.dcount = 2
        self.dexport = openpyxl.Workbook()
        self.dxlsxconfig = None
        self.dconfig = self.data_config()

    # 创建或者读取配置文件
    def init_config(self,path_config):
        print("init_config: " + path_config)
        if os.path.exists(path_config):
            self.dxlsxconfig = openpyxl.load_workbook(path_config)
            dsheet = self.dxlsxconfig["CalcSum"]
            self.dconfig.srcpath = dsheet.cell(2,2).value
            self.dconfig.dstpath = dsheet.cell(3,2).value

            if self.dconfig.dstpath == None:
                self.dconfig.dstpath = ""
            if self.dconfig.srcpath == None:
                self.dconfig.srcpath = ""
            return True
        else:
            self.dxlsxconfig = openpyxl.Workbook()
            dsheet = self.dxlsxconfig.active
            dsheet.title = "CalcSum"
            dsheet.cell(1,1,"Variable")
            dsheet.cell(1,2,"Value")
            dsheet.cell(1,3,"Note")

            dsheet.cell(2,1,"SourcePath")
            dsheet.cell(2,3,"数据来源不能为空")

            dsheet.cell(3,1,"ResultPath")
            dsheet.cell(3,3,"导出结果为空则在当前路径否则会指定路径")
            self.dxlsxconfig.save(path_config)
            return False



    # 写入sheet的计算结果
    def write_ret(self,ct : data_ret):
        print("===========")
        print(ct.sheet_name)
        print(ct.err_tips)
        print(ct.ret_sum)
        print(ct.ret_sum_empty)
        print(ct.ret_sum_date)
        print(ct.ret_count_empty)

        dsheet = self.dexport.active
        if ct.err_tips == "":
            dsheet.cell(self.dcount,1,ct.sheet_name)
            dsheet.cell(self.dcount,2,ct.ret_sum)
            dsheet.cell(self.dcount,3,ct.ret_sum_empty)
            dsheet.cell(self.dcount,4,ct.ret_sum_date)
            dsheet.cell(self.dcount,5,ct.ret_count_empty)
            dsheet.cell(self.dcount,6,ct.err_tips)
        else:
            dsheet.cell(self.dcount,1,ct.sheet_name)
            dsheet.cell(self.dcount,6,ct.err_tips)
        self.dcount += 1
            


    # 导出表格文件
    def write_export_xlsx(path_xlsx):
        print("")



    # 传入每页sheet并从中解析数据
    def parse_sheet_data(self, dsheet : openpyxl.worksheet.worksheet.Worksheet):
        print("parse_sheet_data: " + dsheet.title)
        
        errstr = ""
        if dsheet.cell(2,1).value == None:
            errstr = "不存在表头 <收料日期>"
        elif dsheet.cell(2,4).value == None:
            errstr = "不存在表头 <數量>"
        elif dsheet.cell(2,6).value == None:
            errstr = "不存在表头 <出貨日期>"

        stdate = dsheet.cell(2,1).value
        stnum = dsheet.cell(2,4).value
        stoutdate = dsheet.cell(2,6).value


        # err quit     
        if errstr != "":
            retct = self.data_ret()
            retct.err_tips = errstr
            retct.sheet_name = dsheet.title
            return retct
        
        sum_num : float = 0.0
        sum_empty_date = 0.0
        count_empty_date = 0
        index = 3
        while True:
            if dsheet.cell(index,1).value == None:
                for i in range(1,10):
                    if dsheet.cell(index+i,1).value != None:
                        errstr = "存在空值 < {} > < {} >".format(stdate,index)
                        break
                break
            
            val = dsheet.cell(index,4).value
            if val != None:
                try:
                    float(dsheet.cell(index,4).value)
                except:
                    errstr = "存在空值 < {} > < {} >".format(stnum,index)
                    break
                num = dsheet.cell(index,4).value
                sum_num += num
                if dsheet.cell(index,6).value == None:
                    sum_empty_date += num
                    count_empty_date += 1

            # next row
            index += 1

        # err quit        
        if errstr != "":
            retct = self.data_ret()
            retct.err_tips = errstr
            retct.sheet_name = dsheet.title
            return retct
        
        ret_sum = sum_num
        ret_sum_empty = sum_empty_date
        ret_sum_date = sum_num - sum_empty_date
        ret_count_empty = count_empty_date

        retct = self.data_ret()
        retct.ret_sum = ret_sum
        retct.ret_sum_empty = ret_sum_empty
        retct.ret_sum_date = ret_sum_date
        retct.ret_count_empty = ret_count_empty
        retct.err_tips = errstr
        retct.sheet_name = dsheet.title
        return retct



    # 打开xlsx文件并分出每个sheet页
    def open_xlsx_file(self,path_xlsx):
        print("open_xlsx_file: " + path_xlsx)

        dxlsx = openpyxl.load_workbook(path_xlsx)
        lssheet = dxlsx.sheetnames
        for name in lssheet:
            nowsheet = dxlsx[name]
            ret = self.parse_sheet_data(nowsheet)
            self.write_ret(ret)



    # 创建并保存导出结果表格
    def create_xlsx(self,path_xlsx):
        print("create_xlsx: " + path_xlsx)

        dsheet = self.dexport.active
        dsheet.cell(1,1,"表头")
        dsheet.cell(1,2,"收货总数量")
        dsheet.cell(1,3,"已出货总数量")
        dsheet.cell(1,4,"未出货总数量")
        dsheet.cell(1,5,"未出货总箱数")
        dsheet.cell(1,6,"错误提示")
        self.dexport.save(path_xlsx)



def write_err(path,serr):
    with open(path,"w",encoding="utf-8") as file:
        file.write("错误提示：\n")
        file.write(serr)
        file.close()

# main
if __name__ == "__main__":
    path_config = "./config/config.xlsx"
    path_err = "./config/error.txt"
    path_default = "./config/out.xlsx"

    if os.path.exists(path_err):
        os.remove(path_err)
    if os.path.exists(path_default):
        os.remove(path_default)
    
    a = xlsx_sum_a()
    if a.init_config(path_config) == False:
        write_err(path_err,"配置文件不存在，经重新生成 < {} >".format(path_config))
        sys.exit(0)
    
    if os.path.exists(a.dconfig.srcpath) == False:
        write_err(path_err,"数据源文件路径不存在 < {} >".format(a.dconfig.srcpath))
        sys.exit(0)

    if a.dconfig.dstpath == "":
        a.dconfig.dstpath = path_default

    a.open_xlsx_file(a.dconfig.srcpath)
    a.create_xlsx(a.dconfig.dstpath)
    print("== end ==")

